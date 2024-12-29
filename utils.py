from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from dataclasses import make_dataclass


def xlref(row, column, zero_indexed=True):
    """
    openpyxl helper to generate Excel cell references.

    Args:
        row (int): Row index (zero-indexed or one-indexed based on zero_indexed).
        column (int): Column index (zero-indexed or one-indexed based on zero_indexed).
        zero_indexed (bool, optional): Whether the row and column indices are zero-indexed.
                                      Defaults to True.

    Returns:
        str: Excel cell reference (e.g., "A1", "B2").
    """
    if zero_indexed:
        row += 1
        column += 1
    return get_column_letter(column) + str(row)


def custom_layout_sheet(sheet):
    """
    Openpyxl helper to apply a custom layout to a worksheet.

    This function:
        - Freezes the first row.
        - Adds a filter to the entire sheet.
        - Auto-sizes columns based on cell content.
        - Makes the first row bold.
    """
    for i in range(0, sheet.max_column + 1):
        sheet.freeze_panes = xlref(1, i)

    sheet.auto_filter.ref = sheet.dimensions

    for letter in range(1, sheet.max_column + 1):
        column_letter = get_column_letter(letter)
        max_width = 0
        for cell in sheet[column_letter]:
            if cell.value:
                max_width = max(max_width, len(str(cell.value)))
        sheet.column_dimensions[column_letter].width = (max_width + 1) * 1.25

    for cell in sheet[1]:  # Make first row bold
        cell.font = Font(bold=True)


def read_excel_tab(wb, sheet_name, fields):

    """
    Reads data from an Excel sheet and returns a list of data class objects.

    Args:
        wb (openpyxl.Workbook): The Excel workbook object.
        sheet_name (str): The name of the sheet to read.
        fields (list): A list of tuples, where the first element of each tuple is the
                      column name in the sheet and the second element is the
                      corresponding attribute name for the data class.

    Returns:
        list: A list of data class objects containing the extracted data.
    """
    sheet = wb[sheet_name]

    col_name_to_col_index = {}
    for index, column in enumerate(sheet.iter_cols(1, sheet.max_column)):
        if column[0].value:
            col_name_to_col_index[column[0].value.strip()] = index

    header_names = [element[0] for element in fields]
    attr_names = [element[1] for element in fields]

    data_class = make_dataclass('DataClass', attr_names)
    data = []
    for row in sheet.iter_rows(min_row=2):  # Skip the header row
        table_row = [str(cell.value).strip() if cell.value is not None else None
                     for cell in row]
        row_data = [table_row[col_name_to_col_index[header_name]] for header_name in header_names]
        if row_data:  # Skip empty rows
            data.append(data_class(*row_data))
    return data